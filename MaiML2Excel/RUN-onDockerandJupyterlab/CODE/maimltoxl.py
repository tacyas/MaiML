import sys, os
import argparse
from openpyxl import Workbook, load_workbook
import xmltodict, json

from staticClass import maimlelement, filepath, headerlist, fill
from namespace import defaultNS
from Createxldata.CODE.tmp.maimldicttotxt import READMAIMLTOTXT

### Log #################################################
from logging import getLogger, config
with open(filepath.codedir+'LOG/log_config.json', 'r') as f:
    log_conf = json.load(f)
    # logging設定
    config.dictConfig(log_conf)
loggerI = getLogger('maimltoxlI')
loggerD = getLogger('maimltoxlD')

## open&read excel file #################################################
## not called
def readxl(path):
    if path == '':
        path = filepath.input_dir+'newmaimldata.xlsx'
    wb = load_workbook(path)
    return wb

## open&read maiml file #################################################
def readmaiml(filepath):
    with open(filepath, 'r') as inF:
        maiml_dic = xmltodict.parse(inF.read(), process_namespaces=True, namespaces=defaultNS.namespaces, encoding='UTF-8')
        return maiml_dic

### A1~XFDまでのcellのidを取得する #################################################
## not called
def getcellid(x,y):
    #A:x=0
    x=x-1
    code = ord('A')
    # columnの２桁以上
    tf = 1
    xhcode = ord('A')
    ch = ''
    while True:
        loggerD.debug('tf=='+tf)
        loggerD.debug('x=='+x)
        if (tf-1)<(x+1) and (x+1)/26<=tf:
            if tf-2 >= 0:
                xhnum = xhcode + tf-2
                ch = chr(xhnum)
                loggerD.debug(ch)
            #if ch == 'Z':
            break
        tf += 1

    thiscode = ch + chr(code + x%26) + str(y)
    loggerD.debug(thiscode)
    #print(thiscode)
    return thiscode



### headerをエクセルシートに書き出す #################################################
def addHeader(ws):
    for cellid, header in headerlist.items():
        ws[cellid+'1'] = header
        ws[cellid+'1'].fill = fill
    return ws

### 汎用データコンテナのデータをエクセルシートに書き出す ###################################
def addcellgeneralcontainer(generallist, selectkey, g, ws, y, oyahie='', thishie=''):
    #hie = 1  # 階層を表す ex)1,1-1,2-1-3,9-5
    #y=2  # エクセルのy(1~),行数
    if isinstance(generallist,list):
        pass
    else:
        generallist = [generallist]
    for generaldict in generallist:
        #hierarchy,element
        if oyahie == '':
            hievalue = str(thishie)
        else:
            hievalue = str(oyahie) + '-' + str(thishie)
        if selectkey==[] or generaldict[maimlelement.keyd] in selectkey:
            x=1 # エクセルのx(A~/1~),列数
            hiecell = ws.cell(row=y, column=x)
            hiecell.value = hievalue
            elecell = ws.cell(row=y, column=2)
            if g == 'p':
                elecell.value =maimlelement.property
            if g == 'c':
                elecell.value =maimlelement.content
            if g == 'u':
                elecell.value =maimlelement.uncertainty
            #headerの値に合わせてmaimlデータを取得セルデータに格納
            for cellid, header in headerlist.items():
                if header in generaldict.keys():
                    thiscellid = cellid+str(y)
                    ws[thiscellid] = generaldict[header]
            y+=1         # エクセルの行を一つ進める
        
        #多階層の場合
        hie2 = 1
        if maimlelement.property in generaldict.keys():
            propertylist2 = generaldict[maimlelement.property]
            #print('propertylist2:::',propertylist2)
            propertylist2, ws, y, hie2 = addcellgeneralcontainer(propertylist2, selectkey, 'p', ws, y, oyahie=hievalue, thishie=hie2)
        if maimlelement.content in generaldict.keys():
            contentlist2 = generaldict[maimlelement.content]
            #print('contentlist:::',contentlist2)
            contentlist2, ws, y, hie2 = addcellgeneralcontainer(contentlist2, selectkey, 'c', ws, y, oyahie=hievalue, thishie=hie2)
        if maimlelement.uncertainty in generaldict.keys():
            uncertaintylist2 = generaldict[maimlelement.uncertainty]
            #print('uncertaintylist2:::',uncertaintylist2)
            uncertaintylist2, ws, y, hie2 = addcellgeneralcontainer(uncertaintylist2, selectkey, 'u', ws, y, oyahie=hievalue, thishie=hie2)
        thishie += 1
    return generallist,ws,y,thishie


### MaiMLファイルから抽出したresult要素が持つ汎用データコンテナの値をエクセルに書き出す ##########
def createcsv( resultId=[], selectkey=[], maiml_file_path = '', xl_file_path=filepath.output_dir+'output.xlsx'):
    loggerD.debug('createcsv')
    # maiml file --> resultlist
    maimldict = {}
    if maiml_file_path == '':
        loggerI.error('input maiml filepath is null.')
        sys.exit()
    elif not os.path.isfile(maiml_file_path):
        loggerI.error('input maiml filepath is not exist.')
        sys.exit()
    else:
        maimldict = readmaiml(maiml_file_path)
    resultlist = []
    if maimlelement.data in maimldict[maimlelement.maiml].keys() and maimlelement.results in maimldict[maimlelement.maiml][maimlelement.data].keys() and maimlelement.result in maimldict[maimlelement.maiml][maimlelement.data][maimlelement.results].keys():
        resultlist = maimldict[maimlelement.maiml][maimlelement.data][maimlelement.results][maimlelement.result]
    loggerD.debug(resultlist)

    # create & save workbook
    wb = Workbook()
    wb.save(xl_file_path)
    #wb = load_workbook(xl_file_path)
    if isinstance(resultlist,list):
        pass
    else:
        resultlist = [resultlist]
    for resultdict in resultlist:
        loggerD.debug('resultId:{0}'.format(resultId))
        if resultId == [] or resultdict[maimlelement.idd] in resultId:
            if any(key in resultdict for key in (maimlelement.property, maimlelement.content, maimlelement.uncertainty)):
                # ws = result-id
                ws_name = resultdict[maimlelement.idd]
                loggerD.debug('create ws. ws-name is {0}'.format(ws_name))
                # create worksheet
                ws = wb.create_sheet(ws_name)
                # add header
                ws = addHeader(ws)
                hie = 1  # 階層を表す ex)1,1-1,2-1-3,9-5
                y=2  # エクセルのy(1~),行数
                #property
                if maimlelement.property in resultdict.keys():
                    propertylist = resultdict[maimlelement.property]
                    #loggerD.debug('propertylist:::{0}'.format(propertylist))
                    propertylist, ws, y, hie = addcellgeneralcontainer(propertylist, selectkey, 'p', ws, y, thishie=hie)
                #content
                if maimlelement.content in resultdict.keys():
                    contentlist = resultdict[maimlelement.content]
                    #loggerD.debug('contentlist:::{0}'.format(contentlist))
                    contentlist, ws, y, hie = addcellgeneralcontainer(contentlist, selectkey, 'c', ws, y, thishie=hie)
                #uncertainty
                if maimlelement.uncertainty in resultdict.keys():
                    uncertaintylist = resultdict[maimlelement.uncertainty]
                    #loggerD.debug('uncertaintylist:::{0}'.format(uncertaintylist))
                    uncertaintylist, ws, y, hie = addcellgeneralcontainer(uncertaintylist, selectkey, 'u', ws, y, thishie=hie)
                wb.save(xl_file_path)
    return

### open&read input.json #############################################################################
## 'maiml_file_name' : Required
## 'xl_file_name' and 'resultId' and 'selectkey' : not Required
############################
def openinputjson():
    jsonfile = open(filepath.input_dir+'input.json', 'r')
    inputlist = json.load(jsonfile)
    loggerD.debug('input data: {0}'.format(inputlist))
    if 'maiml_file_name' not in inputlist or inputlist['maiml_file_name'] == '':
        loggerI.error('"input.json" file data is incorrect.')
        loggerI.error('"maiml_file_name" is required.')
        sys.exit()
    else:
        inputlist['maiml_file_name'] = filepath.input_dir + inputlist['maiml_file_name']
    if 'xl_file_name' not in inputlist or inputlist['xl_file_name'] == "":
        output_filepath = filepath.output_dir+'output.xlsx'
        loggerI.info('use default outout filepath : "{0}".'.format(output_filepath))
        inputlist['xl_file_name'] = output_filepath
    else:
        inputlist['xl_file_name'] = filepath.output_dir + inputlist['xl_file_name']
    if 'resultId' not in inputlist:
        inputlist['resultId'] = []
    if 'selectkey' not in inputlist:
        inputlist['selectkey'] = []
    return inputlist['maiml_file_name'], inputlist['xl_file_name'], inputlist['resultId'], inputlist['selectkey']



## for test ################################################
def testmethod(args):
    input_filepath = ''
    if args.json :
        loggerI.info('use json file. Input json-filepath="{0}"'.format(filepath.input_dir+'input.json'))
        input_filepath,output_filepath,resultId,selectkey = openinputjson()
    else:
        input_filepath = args.maiml
        if args.xl == '':
            output_filepath = filepath.output_dir+'output.xlsx'
            loggerI.info('use default outout filepath : "/DATA/TMP/output.xlsx".')
        else:
            output_filepath = args.xl
        if args.selectid != '':
            resultID = args.selectid
        if args.selectkey != '':
            selectkey = args.selectkey
    input_filepath = filepath.input_dir+'test.maiml'
    output_filepath = filepath.output_dir+'newmaimldata.xlsx'
    ###引数チェック
    ## OK
    input_filepath = filepath.input_dir+'test1.maiml'
    output_filepath = filepath.output_dir+'newmaimldata1.xlsx'
    createcsv(maiml_file_path=input_filepath, xl_file_path=output_filepath)
    ## OK
    input_filepath = filepath.input_dir+'test1.maiml'
    #os.makedirs(os.path.dirname('./tmp/'), exist_ok=True)
    createcsv(maiml_file_path=input_filepath)
    ## OK
    input_filepath = filepath.input_dir+'test1.maiml'
    output_filepath = filepath.output_dir+'newmaimldata2.xlsx'
    createcsv(resultId=['juyosample_RT_ver1-1_instance','2juyosample_RT_ver1-1_instance'],maiml_file_path=input_filepath,xl_file_path=output_filepath)
    ## OK
    input_filepath = filepath.input_dir+'test1.maiml'
    output_filepath = filepath.output_dir+'newmaimldata3.xlsx'
    createcsv(selectkey='tiff:StripOffsets',maiml_file_path=input_filepath,xl_file_path=output_filepath)
    ## OK
    input_filepath = filepath.input_dir+'test1.maiml'
    output_filepath= filepath.output_dir+'newmaimldata4.xlsx'
    createcsv(selectkey=['tiff:StripOffsets', 'tiff:SMinSampleValue','tiff:Compression'],maiml_file_path=input_filepath,xl_file_path=output_filepath)
    
    ## NG:message:input maiml filepath is not exist.
    #input_filepath = input_dir+'testng.maiml'
    #createcsv(maiml_file_path=input_filepath,xl_file_path=output_filepath)
    ## NG:message:input maiml filepath is null.
    #createcsv(xl_file_path=output_filepath)

    ###MaiMLデータチェック###########################################
    # OK : 汎用データが存在しないファイル1:dataコンテンツがない-->空のエクセルファイルを出力
    input_filepath = filepath.input_dir+'test5.maiml'
    output_filepath= filepath.output_dir+'newmaimldata5.xlsx'
    createcsv(maiml_file_path=input_filepath, xl_file_path=output_filepath)
    # OK : 汎用データが存在しないファイル2:resultコンテンツがない-->空のエクセルファイルを出力
    input_filepath = filepath.input_dir+'test6.maiml'
    output_filepath = filepath.output_dir+'newmaimldata6.xlsx'
    createcsv(maiml_file_path=input_filepath, xl_file_path=output_filepath)
    # OK : 汎用データが存在しないファイル3:resultはあるが汎用データコンテンツがない-->空のエクセルファイルを出力
    input_filepath = filepath.input_dir+'test7.maiml'
    output_filepath = filepath.output_dir+'newmaimldata7.xlsx'
    createcsv(maiml_file_path=input_filepath, xl_file_path=output_filepath)
    # OK : 様々なテストデータ1
    input_filepath = filepath.input_dir+'test8.maiml'
    output_filepath= filepath.output_dir+'newmaimldata8.xlsx'
    createcsv(maiml_file_path=input_filepath, xl_file_path=output_filepath)
    # OK : 様々なテストデータ2
    input_filepath = filepath.input_dir+'test9.maiml'
    output_filepath = filepath.output_dir+'newmaimldata9.xlsx'
    createcsv(maiml_file_path=input_filepath, xl_file_path=output_filepath)



### RUN ################################################################################################
###### パラメータ
#######  1 : input file name of maiml data
#######  2 : output file name of excel ( if no-data, use default path "/DATA/OUTPUT/output.xlsx" )
#######  3 : result ID
#######  4 : key data
####### '-t' : テストコード実行
#######  nothing : /DATA/INPUT/input.jsonに記載のデータを取得
################
if __name__ == '__main__':
    args = sys.argv

    ## 引数読み込み
    parser = argparse.ArgumentParser()
    # '-t' > '-j' > '-m'
    parser.add_argument('-j', '--json', action='store_true') # use json file Flag
    parser.add_argument('-m', '--maiml',default='') # input maiml file name
    parser.add_argument('-o', '--xl', default='') # output file name
    parser.add_argument('-si', '--selectid',  nargs="*", default='') # result ID
    parser.add_argument('-sk', '--selectkey',  nargs="*", default='') # general container key
    parser.add_argument('-t', '--test', action='store_true') # tests run Flag

    args = parser.parse_args()
    loggerI.info('run argument:{0}'.format(args))

    # test
    if args.test:
        loggerD.info('test run!')
        testmethod(args)
    else:
        input_filepath = ''
        resultID = []
        selectkey = []
        if args.json:  #if(len(sys.argv) <= 1):
            loggerI.info('Use json file. Input json-filepath="{0}"'.format(filepath.input_dir+'input.json'))
            input_filepath,output_filepath,resultID,selectkey = openinputjson()
        elif args.maiml != '':  #elif(len(args) >= 2):
            input_filepath = filepath.input_dir+args.maiml
            if args.xl == '':
                output_filepath = filepath.output_dir+'output.xlsx'
                loggerI.info('use default outout filepath : "{0}".'.format(output_filepath))
            else:
                output_filepath = filepath.output_dir+args.xl
            if args.selectid != '':
                resultID = args.selectid 
            if args.selectkey != '':
                selectkey = args.selectkey
        else:
            loggerI.error('Arguments is incorrect.')
            sys.exit()
        try:
            loggerI.info('Input filepath="{0}"'.format(input_filepath))
            loggerD.debug('result ID:{0} , key:{1}'.format(resultID, selectkey))
            createcsv(maiml_file_path=input_filepath, xl_file_path=output_filepath, resultId=resultID,selectkey=selectkey)
            loggerI.info('Successful creation of Excel file. Output filepath="{0}"'.format(output_filepath))
        except Exception as e:
            loggerD.error('Fail! {0}'.format(e))
            loggerI.error('Fail! Some error has occured.')