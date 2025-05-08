import sys, os, re, json
from openpyxl import Workbook, load_workbook
import lxml.etree as ET
from staticClass2 import maimlelement, filepath, commandargs, headerlist, headerlistET, fill, encheaderlist


### Log #################################################
from logging import getLogger, config, handlers  ## handlersはconfig内で使用しているので必要
with open(filepath.codedir+'LOG/log_config.json', 'r') as f:
    log_conf = json.load(f)
    ## filehandlerのfilepathの追加
    infologfilename=filepath.codedir+'LOG/INFO.log'
    debuglogfilename=filepath.codedir+'LOG/DEBUG.log'
    data_as_dict = dict(log_conf)
    data_as_dict['handlers']['fileHandler'].update({ "filename": infologfilename})
    data_as_dict['handlers']['fileHandler2'].update({ "filename": debuglogfilename})
    # logging設定
    config.dictConfig(log_conf)
# ロガーに追加
loggerI = getLogger('maimltoxlI')
loggerD = getLogger('maimltoxlD')


### lxmlを使用しXML解析 #################################################
def readmaiml(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        xml = f.read().encode('utf-8')
    root = ET.fromstring(re.sub(rb'xmlns=".*?"', b'', xml, count=1))
    return root

### headerをエクセルシートに書き出す #################################################
def addHeader(ws):
    for cellid, header in headerlist.items():
        ws[cellid+'1'] = header
        ws[cellid+'1'].fill = fill
    return ws

### insertion要素をエクセルシートに書き出す ##################################################
def addinsertioncontent(insertionlist, ws, y, thishie=''):
    if not isinstance(insertionlist,list):
        insertionlist = [insertionlist]
    for insertiontag in insertionlist:
        hashvalue = insertiontag.find(maimlelement.hash).text

        thisline = insertiontag.sourceline
        linecell = ws.cell(row=y, column=1)
        linecell.value = str(thisline)
        hie=1
        # insertion要素は階層と要素名のみ
        hiecell = ws.cell(row=y, column=2)
        hiecell.value = str(thishie)
        elecell = ws.cell(row=y, column=3)
        elecell.value =maimlelement.insertion
        y+=1
        # insertion要素が持つuri要素
        uricell = ws.cell(row=y, column=2)
        uricell.value = str(thishie)+'-'+str(hie)
        uricell = ws.cell(row=y, column=3)
        uricell.value = maimlelement.uri
        urivaluecell = ws.cell(row=y, column=5)
        urivaluecell.value = insertiontag.find(maimlelement.uri).text
        y+=1
        hie+=1
        # insertion要素が持つhash要素
        hashcell = ws.cell(row=y, column=2)
        hashcell.value = str(thishie)+'-'+str(hie)
        hashcell = ws.cell(row=y, column=3)
        hashcell.value = maimlelement.hash
        hashvaluecell = ws.cell(row=y, column=5)
        hashvaluecell.value = hashvalue
        
        y+=1
        # insertion要素が持つuuid要素が存在したら
        if insertiontag.find(maimlelement.uuid) is not None:
            hie+=1
            uuidcell = ws.cell(row=y, column=2)
            uuidcell.value = str(thishie)+'-'+ str(hie)
            uuidcell = ws.cell(row=y, column=3)
            uuidcell.value = maimlelement.uuid
            uuidvaluecell = ws.cell(row=y, column=5)
            uuidvaluecell.value = insertiontag.find(maimlelement.uuid).text
            y+=1
        # insertion要素が持つformat要素が存在したら
        if insertiontag.find(maimlelement.format) is not None:
            hie+=1
            formathiecell = ws.cell(row=y, column=2)
            formathiecell.value = str(thishie)+'-'+str(hie)
            formatcell = ws.cell(row=y, column=3)
            formatcell.value = maimlelement.format
            formatvaluecell = ws.cell(row=y, column=5)
            formatvaluecell.value = insertiontag.find(maimlelement.format).text
            y+=1
        thishie+=1
    return ws, y , thishie

### 汎用データコンテナのデータをエクセルシートに書き出す ###################################
def addcellgeneralcontainer(generallist, selectkey, g, ws, y, oyahie='', thishie=''):
    generallist =  generallist if isinstance(generallist,list) else [generallist]
    for generaltag in generallist:
        #hierarchy,element
        if oyahie == '':
            hievalue = str(thishie)
        else:
            hievalue = str(oyahie) + '-' + str(thishie)
        tagdict = generaltag.attrib
        if selectkey==[] or tagdict[maimlelement.key] in selectkey:
            ## add maimlfile lineNo
            lineNo = generaltag.sourceline
            linecell = ws.cell(row=y, column=1)
            linecell.value = lineNo

            x=2 # エクセルのx(A~/1~),列数
            hiecell = ws.cell(row=y, column=2)
            hiecell.value = hievalue
            elecell = ws.cell(row=y, column=3)
            if g == 'p':
                elecell.value =maimlelement.property
            if g == 'c':
                elecell.value =maimlelement.content
            if g == 'u':
                elecell.value =maimlelement.uncertainty

            #headerの値に合わせてmaimlデータを取得セルデータに格納
            for cellxid, header in headerlistET.items():
                thiscellid = cellxid+str(y)
                if generaltag.get(header) is not None:  ## attribute
                    ws[thiscellid] = generaltag.get(header)
                if generaltag.find(header) is not None:  ## child element
                    ## value > 1
                    if header == maimlelement.value:
                        if isinstance(generaltag.findall(maimlelement.value),list):
                            num = 1
                            for valuedata in generaltag.findall(maimlelement.value):
                                ws[thiscellid] = valuedata.text
                                if num < len(generaltag.findall(maimlelement.value)):
                                    y+=1
                                    thiscellid = cellxid+str(y)
                                num += 1
                        else:
                            ws[thiscellid] = generaltag.find(maimlelement.value).text
                    else:
                        ws[thiscellid] = generaltag.find(header).text
        y+=1         # エクセルの行を一つ進める

        # 汎用データコンテナ内で秘匿化されたコンテンツをもつ場合
        hie2 = 1
        for encheader in encheaderlist: ## case: encrypted data exists
            if generaltag.find(encheader) is not None:
                #'B':'hierarchy',
                thiscellid = 'B' + str(y)
                ws[thiscellid] = hievalue + '-' + str(hie2)
                # 'C':'element',
                thiscellid = 'C' + str(y)
                ws[thiscellid] = encheader
                # 'E':'element data',
                thiscellid = 'E' + str(y)
                ws[thiscellid] = generaltag.find(encheader).text
                hie2+=1
                y+=1
        
        #多階層の場合
        #hie2 = 1
        if  generaltag.find(maimlelement.property) is not None:
            propertylist2 = generaltag.findall(maimlelement.property)
            propertylist2, ws, y, hie2 = addcellgeneralcontainer(propertylist2, selectkey, 'p', ws, y, oyahie=hievalue, thishie=hie2)
        if generaltag.find(maimlelement.content) is not None:
            contentlist2 = generaltag.findall(maimlelement.content)
            contentlist2, ws, y, hie2 = addcellgeneralcontainer(contentlist2, selectkey, 'c', ws, y, oyahie=hievalue, thishie=hie2)
        if generaltag.find(maimlelement.uncertainty) is not None:
            uncertaintylist2 = generaltag.findall(maimlelement.uncertainty)
            uncertaintylist2, ws, y, hie2 = addcellgeneralcontainer(uncertaintylist2, selectkey, 'u', ws, y, oyahie=hievalue, thishie=hie2)
        thishie += 1
    return generallist,ws,y,thishie


### MaiMLファイルから抽出したresult要素が持つ汎用データコンテナの値をエクセルに書き出す ##########
def createcsv( resultId=[], selectkey=[], maiml_file_path = '', xl_file_path=filepath.output_dir+'output.xlsx'):
    loggerD.debug('createcsv method start.')
    # maiml file --> resultlist
    maimldict = {}
    if maiml_file_path == '':
        loggerI.error('input maiml filepath is null.')
        sys.exit()
    elif not os.path.isfile(maiml_file_path):
        loggerI.error('input maiml filepath is not exist.')
        sys.exit()
    else:
        maimldata = readmaiml(maiml_file_path)
    loggerD.info('parsed maiml.:::{0}'.format(maimldata))
    resultlist = []
    ## results >=1
    if maimldata.find(maimlelement.data) is not None and maimldata.find(maimlelement.data).find(maimlelement.results) is not None:
        #line_numbers = {}
        resultslist = maimldata.find(maimlelement.data).findall(maimlelement.results)
        loggerD.debug('user select resultId:{0}'.format(resultId))
        # create & save workbook
        wb = Workbook()
        wb.save(xl_file_path)
        if not isinstance(resultslist, list):
            resultslist = [resultslist]
        for resultstag in resultslist:
            resultshastaglist = resultstag.findall(maimlelement.result) if isinstance(resultstag.findall(maimlelement.result),list) else [resultstag.findall(maimlelement.result)]
            ## material & conditionも出力したい場合は、次２行のコメントを外す
            resultshastaglist += resultstag.findall(maimlelement.material)
            resultshastaglist += resultstag.findall(maimlelement.condition)
            for resulttag in resultshastaglist:
                if resultId == [] or resulttag.get(maimlelement.id) in resultId:
                    ## maiml file lineNo
                    ws_name = resulttag.get(maimlelement.id)
                    ## create worksheet
                    resultlineno = resulttag.sourceline
                    ws = wb.create_sheet(str(resultlineno) + '-' + ws_name[:20])
                    loggerD.debug('create ws. ws-name is {0}'.format(ws.title))
                    ## add header
                    ws = addHeader(ws)

                    hie = 1  # 階層を表す ex)1,1-1,2-1-3,9-5
                    y=2  # エクセルのy(1~),行数

                    ## insertion tag >=0
                    if resulttag.find(maimlelement.insertion) is not None:
                        insertionlist = resulttag.findall(maimlelement.insertion) if isinstance(resulttag.findall(maimlelement.insertion),list) else [resulttag.findall(maimlelement.insertion)]
                        ws, y, hie = addinsertioncontent(insertionlist, ws, y, thishie=hie)
                    
                    hie2 = 1
                    for encheader in encheaderlist: ## case: encrypted data exists
                        if resulttag.find(encheader) is not None:
                            #'B':'hierarchy',
                            thiscellid = 'B' + str(y)
                            ws[thiscellid] = str(hie) + '-' + str(hie2)
                            # 'C':'element',
                            thiscellid = 'C' + str(y)
                            ws[thiscellid] = encheader
                            # 'E':'element data',
                            thiscellid = 'E' + str(y)
                            ws[thiscellid] = resulttag.find(encheader).text
                            hie2+=1
                            y+=1   
                    
                    ## 汎用データコンテナ
                    if resulttag.find(maimlelement.property) is not None:
                        propertytaglist = resulttag.findall(maimlelement.property)
                        propertytaglist, ws, y, hie = addcellgeneralcontainer(propertytaglist, selectkey, 'p', ws, y, thishie=hie)
                    if resulttag.find(maimlelement.content) is not None:
                        contenttaglist = resulttag.findall(maimlelement.content)
                        contenttaglist, ws, y, hie = addcellgeneralcontainer(contenttaglist, selectkey, 'c', ws, y, thishie=hie)
                    if resulttag.find(maimlelement.uncertainty) is not None:
                        uncertaintytaglist = resulttag.findall(maimlelement.uncertainty)
                        uncertaintytaglist, ws, y, hie =  addcellgeneralcontainer(uncertaintytaglist, selectkey, 'u', ws, y, thishie=hie)
                    
                    wb.save(xl_file_path)

        ## デフォルトで作成されたシートを削除
        for ws in wb.worksheets:
            if ws.title == 'Sheet':
                wb.remove(ws)
                wb.save(xl_file_path)  
    return  

### MaiMLファイルから抽出したprotocol層が持つ汎用データコンテナの値をエクセルに書き出す ##########
def createtempcsv( templateId=[], selectkey=[], maiml_file_path = '', xl_file_path=filepath.output_dir+'templates.xlsx'):
    loggerD.debug('createcsv method start.')
    # maiml file --> resultlist
    maimldict = {}
    if maiml_file_path == '':
        loggerI.error('input maiml filepath is null.')
        sys.exit()
    elif not os.path.isfile(maiml_file_path):
        loggerI.error('input maiml filepath is not exist.')
        sys.exit()
    else:
        maimldata = readmaiml(maiml_file_path)
    loggerD.info('parsed maiml.:::{0}'.format(maimldata))
    templateslist = []

    if maimldata.find(maimlelement.protocol) is not None :
        templateslist += maimldata.find(maimlelement.protocol).findall(maimlelement.materialTemplate)
        templateslist += maimldata.find(maimlelement.protocol).findall(maimlelement.conditionTemplate)
        templateslist += maimldata.find(maimlelement.protocol).findall(maimlelement.resultTemplate)
    if maimldata.find(maimlelement.protocol) is not None and maimldata.find(maimlelement.protocol).findall(maimlelement.method) is not None:
        methodlist = maimldata.find(maimlelement.protocol).findall(maimlelement.method)
        if not isinstance(methodlist, list):
            methodlist = [methodlist]
        for methodtag in methodlist:
            templateslist += methodtag.findall(maimlelement.materialTemplate)
            templateslist += methodtag.findall(maimlelement.conditionTemplate)
            templateslist += methodtag.findall(maimlelement.resultTemplate)
            programlist = methodtag.findall(maimlelement.program)
            if not isinstance(programlist, list):
                programlist = [programlist]
            for programtag in programlist:
                templateslist += programtag.findall(maimlelement.materialTemplate)
                templateslist += programtag.findall(maimlelement.conditionTemplate)
                templateslist += programtag.findall(maimlelement.resultTemplate)
                
        # create & save workbook
        wb = Workbook()
        wb.save(xl_file_path)
        if not isinstance(templateslist, list):
            templateslist = [templateslist]
        for templatestag in templateslist:
            if templateId == [] or templatestag.get(maimlelement.id) in templateId:
                ## maiml file lineNo
                ws_name = templatestag.get(maimlelement.id)
                ## create worksheet
                templateslineno = templatestag.sourceline
                ws = wb.create_sheet(str(templateslineno) + '-' + ws_name[:20])
                loggerD.debug('create ws. ws-name is {0}'.format(ws.title))
                ## add header
                ws = addHeader(ws)

                hie = 1  # 階層を表す ex)1,1-1,2-1-3,9-5
                y=2  # エクセルのy(1~),行数

                hie2 = 1
                for encheader in encheaderlist: ## case: encrypted data exists
                    if templatestag.find(encheader) is not None:
                        #'B':'hierarchy',
                        thiscellid = 'B' + str(y)
                        ws[thiscellid] = str(hie) + '-' + str(hie2)
                        # 'C':'element',
                        thiscellid = 'C' + str(y)
                        ws[thiscellid] = encheader
                        # 'E':'element data',
                        thiscellid = 'E' + str(y)
                        ws[thiscellid] = templatestag.find(encheader).text
                        hie2+=1
                        y+=1   
                
                ## 汎用データコンテナ
                if templatestag.find(maimlelement.property) is not None:
                    propertytaglist = templatestag.findall(maimlelement.property)
                    propertytaglist, ws, y, hie = addcellgeneralcontainer(propertytaglist, selectkey, 'p', ws, y, thishie=hie)
                if templatestag.find(maimlelement.content) is not None:
                    contenttaglist = templatestag.findall(maimlelement.content)
                    contenttaglist, ws, y, hie = addcellgeneralcontainer(contenttaglist, selectkey, 'c', ws, y, thishie=hie)
                if templatestag.find(maimlelement.uncertainty) is not None:
                    uncertaintytaglist = templatestag.findall(maimlelement.uncertainty)
                    uncertaintytaglist, ws, y, hie =  addcellgeneralcontainer(uncertaintytaglist, selectkey, 'u', ws, y, thishie=hie)
                
                wb.save(xl_file_path)

        ## デフォルトで作成されたシートを削除
        for ws in wb.worksheets:
            if ws.title == 'Sheet':
                wb.remove(ws)
                wb.save(xl_file_path)  
    return  

### open&read input.json #############################################################################
## 'maiml_file_name' : Required
## 'xl_file_name' and 'resultId' and 'selectkey' : not Required
############################
def openinputjson():
    try:
        jsonfile = open(filepath.input_dir+'input.json', 'r')
    except Exception as e:
        loggerI.error('"{0}" file not found.'.format(filepath.input_dir+'input.json'))
        sys.exit()
    inputlist = json.load(jsonfile)
    loggerD.debug('input data: {0}'.format(inputlist))
    if 'maiml_file_name' not in inputlist or inputlist['maiml_file_name'] == '':
        loggerI.error('"input.json" file data is incorrect.')
        loggerI.error('"maiml_file_name" is required.')
        sys.exit()
    else:
        inputlist['maiml_file_name'] = filepath.input_dir + inputlist['maiml_file_name']
    if 'xl_file_name' not in inputlist or inputlist['xl_file_name'] == "":
        output_filepath = filepath.output_dir+'instances.xlsx'
        loggerI.info('use default outout filepath : "{0}".'.format(output_filepath))
        inputlist['xl_file_name'] = output_filepath
    else:
        inputlist['xl_file_name'] = filepath.output_dir + inputlist['xl_file_name']
    if 'xl_file_name_temp' not in inputlist or inputlist['xl_file_name_temp'] == "":
        output_filepath = filepath.output_dir+'templates.xlsx'
        loggerI.info('use default outout filepath : "{0}".'.format(output_filepath))
        inputlist['xl_file_name_temp'] = output_filepath
    else:
        inputlist['xl_file_name_temp'] = filepath.output_dir + inputlist['xl_file_name_temp']
    if 'resultId' not in inputlist:
        inputlist['resultId'] = []
    if 'selectkey' not in inputlist:
        inputlist['selectkey'] = []
    return inputlist['maiml_file_name'], inputlist['xl_file_name'], inputlist['resultId'], inputlist['selectkey'], inputlist['xl_file_name_temp']


### RUN ################################################################################################
###### パラメータ
#######  -j : use json file 
#######  -m : input file name of maiml data
#######  -o : output file name of excel ( if no-data, use default path "/DATA/OUTPUT/instances.xlsx" )
#######  -si : result ID
#######  -sk : key data
################
if __name__ == '__main__':
    args = commandargs.parser.parse_args()
    loggerI.info('run argument:{0}'.format(args))
  
    input_filepath = ''
    resultID = []
    selectkey = []
    if args.json: # use json
        loggerI.info('use json file. Input json-filepath="{0}"'.format(filepath.input_dir+'input.json'))
        input_filepath,output_filepath,resultID,selectkey,xl_file_name_temp = openinputjson()
    elif args.maiml != '': # use command args
        input_filepath = filepath.input_dir+args.maiml
        if args.xl == '':
            output_filepath = filepath.output_dir+'output.xlsx'
            loggerI.info('use default outout filepath : "{0}".'.format(output_filepath))
        else:
            output_filepath = filepath.output_dir+args.xl
            loggerI.info('use outout filepath : "{0}".'.format(output_filepath))
        if args.selectid != '':
            resultID = args.selectid 
        if args.selectkey != '':
            selectkey = args.selectkey
    else:
        loggerI.error('Arguments is incorrect.')
        sys.exit()
    try:
        loggerI.info('Input filepath="{0}"'.format(input_filepath))
        loggerD.debug('result ID:{0} , key:{1}'.format(resultID, selectkey))
        createcsv(maiml_file_path=input_filepath, xl_file_path=output_filepath, resultId=resultID,selectkey=selectkey)
        createtempcsv(maiml_file_path=input_filepath, xl_file_path=xl_file_name_temp, selectkey=selectkey)
        loggerI.info('Successful creation of Excel file. Output filepath="{0}"'.format(output_filepath))
    except Exception as e:
        loggerD.error('Fail! {0}'.format(e))
        loggerI.error('Fail! Some error has occured.')